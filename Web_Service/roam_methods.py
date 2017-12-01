
from pandas import read_csv,ExcelFile,ExcelWriter
from numpy import where,inf,nan,arange
from openpyxl import load_workbook
import os
import matplotlib.pyplot as plt


def find_nearest(array_1,array_2):
    indexes_tochange = []
    for elements in array_2:
        try:
            k = max(filter(lambda x: x<elements,array_1))
            indexes_tochange.append(k)
        except:
            pass
    return indexes_tochange


def get_dataFrame(path):
    wlan_steering_path = path[:-4] + '_steering_wlan.txt'
    col_names = ['timestamp','eventtype','eventname','senderMac','clientMac','targetMac','currentCost','currentRSSI','targetCost','targetRSSI','inferance','sequence']
    os.system("sed -r '/steering|wlan/!d' "+path+" > "+wlan_steering_path)
    df = read_csv(wlan_steering_path, error_bad_lines=False, delim_whitespace=True, header=None, low_memory=False, names=col_names)

    df['clientMac'] = where(df['eventname'] == 'registered' , df['senderMac'], df['clientMac'])
    df['clientMac'] = where(df['eventname'] == 'expired', df['senderMac'], df['clientMac'])
    df['clientMac'] = where(df['eventname']=='remove_rssi_threshold_2_4_ghz', df['senderMac'],df['clientMac'])
    df['clientMac'] = where(df['eventname'] == 'dynamic_untouchable_add', df['senderMac'], df['clientMac'])
    df['clientMac'] = where(df['eventname'] == 'dynamic_untouchable_del', df['senderMac'], df['clientMac'])

    df['clientMac'] = df['clientMac'].str.lower()
    df['senderMac'] = df['senderMac'].str.lower()
    df['targetMac'] = df['targetMac'].str.lower()

    df['targetRSSI'] = df['targetRSSI'].fillna(inf)
    df['currentRSSI'] = df['currentRSSI'].fillna(inf)
    df = df[df.timestamp != nan]
    df['timestamp'] = df['timestamp'].str[1:-1]
    df['timestamp'] = df['timestamp'].astype(dtype=int)
    df = df.sort_values(by=['timestamp','eventname'],ascending=[True,False])
    df = df.reset_index(drop=True)

    return df


def write_xlsx(df,xlsx_file,RSSI_compare,window_time):


    xlsx_df = df
    work = ExcelWriter(xlsx_file, engine='openpyxl')

    for client_mac in set(xlsx_df['clientMac']):
        df_client = xlsx_df[xlsx_df.clientMac == client_mac]
        df_client = df_client.loc[:,['timestamp','eventtype', 'eventname', 'senderMac', 'targetMac', 'currentCost', 'currentRSSI','targetCost', 'targetRSSI', 'sequence']]
        df_client = df_client.sort_values(by=['timestamp']).reset_index(drop=True)


        roams_index_all = list(where(df_client.eventname == 'guided_roam')[0])
        guide_suc_assoc_all = list(where(df_client.eventname == 'guide_success_associate')[0])
        kick_success_all = list(where(df_client.eventname == 'kick_success_associate')[0])
        bsstrans_all = list(where(df_client.eventname == 'guided_roam_bsstrans')[0])
        bss_accepted_all = list(where(df_client.eventname == 'guided_roam_bsstrans_accepted')[0])
        bss_rejected_all = list(where(df_client.eventname == 'guided_roam_bsstrans_rejected')[0])
        bss_unsupported_all = list(where(df_client.eventname == 'guided_roam_bsstrans_unsupported')[0])
        remove_rssi_threshold_all = list(where(df_client.eventname == 'remove_rssi_threshold_2_4_ghz')[0])
        dyn_unt_add_all = list(where(df_client.eventname == 'dynamic_untouchable_add')[0])
        dyn_unt_add_del = list(where(df_client.eventname == 'dynamic_untouchable_del')[0])



        desired_index_for_guide = find_nearest(roams_index_all, guide_suc_assoc_all)
        desired_index_for_kick = find_nearest(roams_index_all, kick_success_all)
        desired_index_for_bsstrans = find_nearest(roams_index_all, bsstrans_all)
        desired_index_for_bss_accepted = find_nearest(roams_index_all, bss_accepted_all)
        desired_index_for_bss_rejected = find_nearest(roams_index_all, bss_rejected_all)
        desired_index_for_bss_unsupported = find_nearest(roams_index_all, bss_unsupported_all)
        desired_index_for_remove = find_nearest(roams_index_all,remove_rssi_threshold_all)
        desired_index_for_dyn_add = find_nearest(roams_index_all,dyn_unt_add_all)
        desired_index_for_dyn_del = find_nearest(roams_index_all,dyn_unt_add_del)


        del remove_rssi_threshold_all[:len(remove_rssi_threshold_all)-len(desired_index_for_remove)]
        del dyn_unt_add_all[:len(dyn_unt_add_all)-len(desired_index_for_dyn_add)]
        del dyn_unt_add_del[:len(dyn_unt_add_del)-len(desired_index_for_dyn_del)]
        del bss_unsupported_all[:len(bss_unsupported_all)-len(desired_index_for_bss_unsupported)]
        del bss_rejected_all[:len(bss_rejected_all)-len(desired_index_for_bss_rejected)]
        del bss_accepted_all[:len(bss_accepted_all)-len(desired_index_for_bss_accepted)]
        del bsstrans_all[:len(bsstrans_all)-len(desired_index_for_bsstrans)]
        del kick_success_all[:len(kick_success_all)-len(desired_index_for_kick)]
        del guide_suc_assoc_all[:len(guide_suc_assoc_all)-len(desired_index_for_guide)]


        for w in range(len(desired_index_for_remove)):
            df_client['currentRSSI'][remove_rssi_threshold_all[w]] = df_client['currentRSSI'][desired_index_for_remove[w]]
        for q in range(len(desired_index_for_dyn_add)):
            df_client['currentRSSI'][dyn_unt_add_all[q]] = df_client['currentRSSI'][desired_index_for_dyn_add[q]]
        for e in range(len(desired_index_for_dyn_del)):
            df_client['currentRSSI'][dyn_unt_add_del[e]] = df_client['currentRSSI'][desired_index_for_dyn_del[e]]
        for k in range(len(desired_index_for_guide)):
            df_client['currentRSSI'][guide_suc_assoc_all[k]] = df_client['currentRSSI'][desired_index_for_guide[k]]
        for l in range(len(desired_index_for_kick)):
            df_client['currentRSSI'][kick_success_all[l]] = df_client['currentRSSI'][desired_index_for_kick[l]]
        for m in range(len(desired_index_for_bsstrans)):
            df_client['currentRSSI'][bsstrans_all[m]] = df_client['currentRSSI'][desired_index_for_bsstrans[m]]
        for n in range(len(desired_index_for_bss_accepted)):
            df_client['currentRSSI'][bss_accepted_all[n]] = df_client['currentRSSI'][desired_index_for_bss_accepted[n]]
        for p in range(len(desired_index_for_bss_rejected)):
            df_client['currentRSSI'][bss_rejected_all[p]] = df_client['currentRSSI'][desired_index_for_bss_rejected[p]]
        for r in range(len(desired_index_for_bss_unsupported)):
            df_client['currentRSSI'][bss_unsupported_all[r]] = df_client['currentRSSI'][desired_index_for_bss_unsupported[r]]


        df_client = df_client[df_client.currentRSSI > RSSI_compare]
        df_client = df_client.reset_index(drop=True)

        eventname = df_client['eventname']
        timestamp = df_client['timestamp']
        senderMac = df_client['senderMac']

        guided_roam_indices = list(where(eventname == 'guided_roam')[0])

        if len(guided_roam_indices) > 0:
            start_index_of_roam = guided_roam_indices[0]
            startrow = 0
            window_end_indice = guided_roam_indices[0]

            for k in range(len(guided_roam_indices) - 1):
                window_start_indice = start_index_of_roam
                if senderMac[guided_roam_indices[k]] == senderMac[guided_roam_indices[k+1]]:
                    if timestamp[guided_roam_indices[k+1]] - timestamp[guided_roam_indices[k]] > window_time:
                        window_end_indice = guided_roam_indices[k+1]
                        start_index_of_roam = guided_roam_indices[k+1]
                        sub_df = df_client[window_start_indice:window_end_indice]
                        sub_df.to_excel(work, client_mac.replace(':', ''), startrow=startrow, index=False, header=True)
                        startrow += len(sub_df['timestamp']) + 3

                if senderMac[guided_roam_indices[k]] != senderMac[guided_roam_indices[k+1]]:
                    window_end_indice = guided_roam_indices[k+1]
                    start_index_of_roam = guided_roam_indices[k+1]
                    sub_df = df_client[window_start_indice:window_end_indice]
                    sub_df.to_excel(work, client_mac.replace(':', ''), startrow=startrow, index=False, header=True)
                    startrow += len(sub_df['timestamp']) + 3

            dfc_end = df_client[window_end_indice:]
            dfc_end.to_excel(work, client_mac.replace(':', ''), startrow=startrow, index=False, header=True)

    try:
        work.save()
    except Exception as e:
        print (e)
        pass



"""def getstatistics():
    RSSI_compare = float(Entry_rssi.get())
    window_time = int(Entry_w_t.get())
    path = Entry_path.get()
    xlsx_file = path[:-4] + '_' + str(window_time)+'_'+ str(int(-1*RSSI_compare)) + ".xlsx"
    print ("---------")
    wb = ExcelFile(xlsx_file)
    sheetnames = wb.sheet_names
    print ("loading workbook")
    wb_insert = load_workbook(xlsx_file)
    print ("workbook loaded")

    if 'Statistics' in sheetnames:
        print("'Statistics sheet already exists. It will be updated.")
        print("------------------------------------")
        statistics_sheet = wb_insert.get_sheet_by_name('Statistics')
        sheetnames.remove('Statistics')
    else:
        wb_insert.create_sheet('Statistics',len(sheetnames))
        statistics_sheet = wb_insert.get_sheet_by_name('Statistics')

    row_client_start = 1

    for clients in sheetnames:
        print(clients)
        #sheet = wb_insert.get_sheet_by_name(clients)
        #highest_row = sheet.get_highest_row()


        df = wb.parse(sheetname=clients,header=None)
        timestamp = df[df.columns[0]]
        timestamp_indexes = list(where(timestamp == 'timestamp')[0])


        num_of_guide_success_associate_11v = 0
        num_of_guide_success_associate_wh_11v = 0
        num_of_kick_success_associate = 0
        num_of_guided_send = 0
        num_of_guided_11v = 0
        num_of_guided_send_fail = 0
        num_of_guided_send_lost = 0
        num_of_11v_window = 0
        num_of_non_11v_window = 0

        for h in range(len(timestamp_indexes)):
            if h != len(timestamp_indexes) - 1:
                window_frame = df[timestamp_indexes[h]:timestamp_indexes[h + 1]]
            else:
                window_frame = df[timestamp_indexes[h]:]

            w_events = list(window_frame[window_frame.columns[2]])
            w_types = list(window_frame[window_frame.columns[1]])
            w_currentrssi = list(window_frame[window_frame.columns[6]])
            send_fail_sequences = []
            send_lost_sequences = []

            for f in range(len(w_events)):
                if w_events[f] == 'guide_success_associate':
                    if 'guided_11v' in w_events:
                        num_of_guide_success_associate_11v += 1
                    else:
                        num_of_guide_success_associate_wh_11v += 1

            if w_events.count('guided_roam') > 1:
                if 'guided_11v' in w_events:
                    num_of_11v_window += 1
                else:
                    num_of_non_11v_window += 1
            else:
                if w_types.count('steering') > 1:
                    if 'guided_11v' in w_events:
                        num_of_11v_window += 1
                    else:
                        num_of_non_11v_window += 1




            for g in range(len(w_events)):
                if w_events[g] == 'kick_success_associate':
                    num_of_kick_success_associate += 1
                if w_events[g] == 'guided_send':
                    num_of_guided_send += 1
                if w_events[g] == 'guided_send_fail':
                    send_fail_sequences.append(w_currentrssi[g])
                if w_events[g] == 'guided_send_lost':
                    send_lost_sequences.append(w_currentrssi[g])
                if w_events[g] == 'guided_11v':
                    num_of_guided_11v += 1


            num_of_guided_send_fail += len(set(send_fail_sequences))
            num_of_guided_send_lost += len(set(send_lost_sequences))


        num_of_guide_success_associate = num_of_guide_success_associate_11v + num_of_guide_success_associate_wh_11v
        window_size = num_of_11v_window + num_of_non_11v_window
        statistics_sheet.cell(row=row_client_start, column=1).value = clients
        if window_size != 0:
            guide_success_ratio = float(num_of_guide_success_associate) / float(window_size)
            statistics_sheet.cell(row= row_client_start, column= 2).value = "guide success ratio"
            statistics_sheet.cell(row=row_client_start+1, column= 2).value = str(round(guide_success_ratio*100,2))+"%" + " (" + str(num_of_guide_success_associate) + "/" + str(window_size) + ")"

        if num_of_11v_window != 0:
            guide_success_ratio_11v = float(num_of_guide_success_associate_11v) / float(num_of_11v_window)
            statistics_sheet.cell(row=row_client_start, column=3).value = "guide success ratio_11v"
            statistics_sheet.cell(row=row_client_start + 1, column=3).value = str(round(guide_success_ratio_11v*100,2))+"%" + " (" + str(num_of_guide_success_associate_11v) + "/" + str(num_of_11v_window) + ")"

        if num_of_non_11v_window != 0:
            guide_success_ratio_non_11v = float(num_of_guide_success_associate_wh_11v) / float(num_of_non_11v_window)
            statistics_sheet.cell(row=row_client_start, column=4).value = "guide success ratio_non_11v"
            statistics_sheet.cell(row=row_client_start + 1, column=4).value = str(round(guide_success_ratio_non_11v*100,2))+"%" + " (" + str(num_of_guide_success_associate_wh_11v) + "/" + str(num_of_non_11v_window) + ")"

        if num_of_guided_11v != 0:
            v11_success_ratio = float(num_of_guide_success_associate_11v) / float(num_of_guided_11v)
            statistics_sheet.cell(row= row_client_start, column= 5).value = "11_v success ratio"
            statistics_sheet.cell(row=row_client_start+1, column= 5).value = str(round(v11_success_ratio*100,2))+"%"+ " (" + str(num_of_guide_success_associate_11v) + "/" + str(num_of_guided_11v) + ")"

        if num_of_guided_send != 0:
            kick_success_ratio = float(num_of_kick_success_associate) / float(num_of_guided_send)
            statistics_sheet.cell(row= row_client_start, column= 6).value = "kick success ratio"
            statistics_sheet.cell(row=row_client_start+1, column= 6).value = str(round(kick_success_ratio*100,2))+"%"+ " (" + str(num_of_kick_success_associate) + "/" + str(num_of_guided_send) + ")"

            kick_fail_loss_ratio = float(max(num_of_guided_send_fail, num_of_guided_send_lost)) / float(num_of_guided_send)
            statistics_sheet.cell(row= row_client_start, column= 7).value = "kick fail-loss ratio"
            statistics_sheet.cell(row=row_client_start+1, column= 7).value = str(round(kick_fail_loss_ratio*100,2))+"%"+ " (" + str(max(num_of_guided_send_fail,num_of_guided_send_lost)) + "/" + str(num_of_guided_send) + ")"

        total_success = num_of_guide_success_associate+num_of_kick_success_associate
        if window_size != 0:
            total_success_ratio = float(total_success)/float(window_size)
            statistics_sheet.cell(row=row_client_start, column=8).value = "total success_ratio"
            statistics_sheet.cell(row=row_client_start+1, column=8).value = str(round(total_success_ratio*100,2))+"%"+" ("+str(total_success)+"/"+str(window_size)+")"

        row_client_start += 3



    wb_insert.save(xlsx_file)


def message():
    messagebox.showinfo("Finished", "Task Finished")


def browse_button():
    filename = askopenfilename(initialdir = "/home/",title = "Select file",filetypes = (("text files","*.txt"),("csv files","*.csv"),("xlsx files","*.xlsx"),("all files","*.*")))
    Entry_path.delete(0, tkinter.END)
    Entry_path.insert(0, filename)


def browse_button_xlsx():
    filename = askopenfilename(initialdir="/home/", title="Select file", filetypes=(("text files", "*.txt"), ("csv files", "*.csv"), ("xlsx files", "*.xlsx"), ("all files", "*.*")))
    Entry_xlsx.delete(0, tkinter.END)
    Entry_xlsx.insert(0, filename)


def histogram():
    xlsx_file = Entry_xlsx.get()
    wb = ExcelFile(xlsx_file)
    sheetnames = wb.sheet_names
    if 'Statistics' in sheetnames:
        sheetnames.remove('Statistics')
    for clients in sheetnames:
        print(clients)
        df = wb.parse(sheetname=clients, header=None)
        timestamp = df[df.columns[0]]
        timestamp_indexes = list(where(timestamp == 'timestamp')[0])
        window_duration_client = []

        for h in range(len(timestamp_indexes)):
            if h != len(timestamp_indexes) - 1:
                window_frame = df[timestamp_indexes[h]:timestamp_indexes[h + 1]]
            else:
                window_frame = df[timestamp_indexes[h]:]

            window_frame = window_frame.reset_index(drop=True)
            eventtype = window_frame[window_frame.columns[1]]
            tstmp = window_frame[window_frame.columns[0]]
            steering_indices = list(where(eventtype == 'steering')[0])
            start_indice = min(steering_indices)
            end_indice = max(steering_indices)
            window_duration_client.append(tstmp[end_indice]-tstmp[start_indice])


        #num_of_windows = len(timestamp_indexes)
        #x = np.linspace(1,num_of_windows,num_of_windows)
        y = window_duration_client

        path_to_save = os.path.dirname(Entry_xlsx.get())+'/histograms/'
        os.makedirs(path_to_save,exist_ok=True)

        bars = plt.bar(x,y)
        for i in range(len(y)):
            if y[i] > 20:
                bars[i].set_color('red')
            else:
                bars[i].set_color('blue')

        plt.xlabel("window number -->")
        plt.ylabel("window duration (sec)")
        plt.title(clients)
        plt.show()



        bin_list = arange(0,max(y)+20,20)
        n,bins,patches = plt.hist(y,bins=bin_list)
        plt.xlabel("window duration (sec)")
        plt.title("steering histogram: "+clients)


        if len(n) > 5:
            alarm_case_number = int(sum(n[5:]))
        else:
            alarm_case_number = 0
        print(str(alarm_case_number)+" cases that have greater window duration  than 100 secs.")
        print("------------------------------------------")
        plt.savefig(path_to_save+clients+'.png')
        plt.close()


"""